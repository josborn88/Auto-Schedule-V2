import selenium, time, openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

dayShift = {'6a':23, '6:30a':23, '7a':23, '7:30a':23, '8a': 23, '8:15a': 24, '8:30a': 25, '8:45a': 26, '9a': 27, '9:15a': 28, '9:30a': 29, '9:45a': 30, '10a': 31, '10:15a': 32, '10:30a': 33, '10:45a': 34, '11a': 35, '11:15a': 36, '11:30a': 37, '11:45a': 38, '12p': 39, '12:15p': 40, '12:30p':41, '12:45p':42, '1p':43, '1:15p': 44, '1:30p': 45, '1:45p': 46, '2p': 47, '2:15p': 48, '2:30p': 49, '2:45p': 50, '3p': 51, '3:15p': 52, '3:30p': 53, '3:45p': 54, '4p': 55, '4:15p': 56, '4:30p': 57, '4:45p': 58, '5p': 59, '5:15p': 60, '5:30p': 61, '5:45p': 61, '6p': 61, '6:15p': 61, '6:30p': 61, '6:45p': 61, '7p': 61, '7:15p': 61, '7:30p':61, '7:45p':61, '8p':61, '8:15p':61, '8:30p': 61, '8:45p':61, '9p': 61, '9:15p':61, '9:30p': 61, '9:45p': 61, '10p': 61, '10:15p': 61, '10:30p':61, '10:45p':61, '11p': 61, '11:15p':61, '11:30p': 61, '11:45p':61, '12a':61}           
eveningShift = {'8a':23, '8:15':23, '8:30a':23, '9a':23, '9:15a':23, '9:30a': 23, '10a':23, '10:30a':23, '11a':23, '11:30a':23, '12p':23, '12:30p':23, '1p':23, '1:30p':23, '2p':23, '2:30p':23, '3p':23, '3:30p':23, '4p':23, '4:30p':23, '5p':23, '5:15p':24, '5:30p':25, '5:45p':26, '6p':27, '6:15p':28, '6:30p':29, '6:45p':30, '7p':31, '7:15p':32, '7:30p': 33, '7:45p':34, '8p':35, '8:15p':36, '8:30p':37, '8:45p':38, '9p':39, '9:15p':40, '9:30p':41, '9:45p':42, '10p':43, '10:15p':44, '10:30p':45, '10:45p':46, '11p':47, '11:15p': 48, '11:30p': 49, '11:45p':50, '12a':51}
dayShiftTimes = list(dayShift.keys())
dayShiftCells =list(dayShift.values())
eveningShiftTimes = list(eveningShift.keys())
eveningShiftCells = list(eveningShift.values())

greyFill = PatternFill(start_color="C0C0C0", end_color = 'C0C0C0', fill_type='solid')

borderGone = Border(left=Side(border_style='none', color='000000'),
				right=Side(border_style='none', color='000000'),
				top=Side(border_style='none', color='000000'),
				bottom=Side(border_style='none', color='000000'))

def scrapeSchedule(weekday):
		
	Workers = ['Aislinn', 'Anthony', 'Caitlin Y', 'Jason', 'Jenn', 'John O', 'Laurie', 'Penny', 'Thaddeus', 'Tyler', 'Alexis', 'Brady', 'Caitlin J', 'Carla', 'Chad B', 'Curt', 'Danielle', 'Eli', 'Emily M', 'Emily W', 'Hannah-J', 'Hayden', 'Jacob', 'Jahmez', 'Jane', 'Jean', 'Jenica', 'Jennifer', 'John K', 'Josh Kap', 'Julia', 'Leah', 'Lia', 'Lucia', 'Mason', 'Matthew', 'Michael', 'Monica', 'Nick', 'Nicole', 'Paul', 'Rachel S', 'Rachel T', 'Rhea', 'Robert', 'Roman', 'Sam', 'Sami']
	WorkerIDs = ['6675421', '5546761', '14234788', '5546773', '5546671', '14234984', '14360396', '5546758', '5574541', '5546749', '32393335', '22277137', '34040088', '5574520', '32112753', '12051076', '32171708', '34040098', '32455620', '32549373', '16564760', '32514875', '32231810', '32393039', '32455646', '5953009', '5574577', '32231874', '8658921', '5574592', '18736600', '32514881', '5574550', '5996605', '33957293', '13187266', '14234858', '5574544', '34040094', '5616373', '31914667', '5574517', '5574559', '33945743', '32231869', '7542140', '34040097', '32298669']
	
	workers = {}
	
	for i in range(0,len(Workers)):   
		try:
			workers[Workers[i]] = browser.find_element_by_css_selector('#uid-'+WorkerIDs[i]+'-'+weekday[0:3]).text
		except NoSuchElementException:
			workers[Workers[i]] = 'Off today'
				
	
	return workers
	
def sortData(weekday):

	eptPoint = []
	eptPointShifts = []
	
	ept = []
	eptShifts =[]
	
	lksPoint =[]
	lksPointShifts = []
	
	lks = []
	lksShifts = []
	
	theaters = []
	theaterShifts = []
	
	animals = []	
	animalsShifts = []
	#sort the data from WhenIWork
	
	
	
	#Strip off extra details,
	
	workersRaw = scrapeSchedule(weekday)
	workers = {key:val for key, val in workersRaw.items() if val != ''}
	workers = {key:val for key, val in workers.items() if val != 'TIME OFF ALL DAY'}
	workers = {key:val for key, val in workers.items() if val != 'Off today'}
	
	rawList = list(workers.keys())
	rawShifts = list(workers.values())
	
	
	for i in range(0,len(rawList)):
		if rawShifts[i].split()[3] == 'FLOOR':
			eptPoint.append(rawList[i])
			eptPointShifts.append(rawShifts[i])
		elif rawShifts[i].split()[3] == 'EPT' or rawShifts[i].split()[3] == 'POISON' or rawShifts[i].split()[3] == 'TRAINING':
			ept.append(rawList[i])
			eptShifts.append(rawShifts[i])
		elif rawShifts[i].split()[3] == 'LKS':
			if 'POINT' in rawShifts[i].split():
				lksPoint.append(rawList[i])
				lksPointShifts.append(rawShifts[i])
			else:
				lks.append(rawList[i])
				lksShifts.append(rawShifts[i])
		elif rawShifts[i].split()[3] == 'LIVING':
			if (len(rawShifts[i].split()) < 6):
				animals.append(rawList[i])
				animalsShifts.append(rawShifts[i])
			elif rawShifts[i].split()[8] == 'EPT' or rawShifts[i].split()[8] == 'POISON':
				ept.append(rawList[i])
				eptShifts.append(rawShifts[i])
			elif rawShifts[i].split()[8] == 'LKS':
				lks.append(rawList[i])
				lksShifts.append(rawShifts[i])
		elif rawShifts[i].split()[3] == 'PLANETARIUM' or rawShifts[i].split()[3] == 'NAT':
			theaters.append(rawList[i])
			theaterShifts.append(rawShifts[i])
	
	
	def cleanShifts(shift):
		for i in range(0,len(shift)):
			shift[i] = shift[i].replace('\nEPT', '')
			shift[i] = shift[i].replace('\nPOISON SHOW', '')
			shift[i] = shift[i].replace('\nLIVING COLLECTIONS', '')
			shift[i] = shift[i].replace('\nLKS', '')
			shift[i] = shift[i].replace('\nFLOOR POINT', '')
			shift[i] = shift[i].replace('\nNAT GEO', '')
			shift[i] = shift[i].replace('\nPLANETARIUM', '')
			shift[i] = shift[i].replace('POINT', '')
			shift[i] = shift[i].replace('TRAINING', '')
		return shift	
	
	eptPointShifts = cleanShifts(eptPointShifts)
	eptShifts = cleanShifts(eptShifts)
	lksPointShifts = cleanShifts(lksPointShifts)
	lksShifts = cleanShifts(lksShifts)
	theatersShifts = cleanShifts(theaterShifts)
	animalsShifts = cleanShifts(animalsShifts)
	
	
	allShifts = [eptPoint, eptPointShifts, ept, eptShifts, theaters, theaterShifts, lks, lksShifts, lksPoint, lksPointShifts, animals, animalsShifts]
	
	return allShifts
	
def addShifts(weekday):

	def isDay(shift):	
		
		day = False
		
		if ('a' in shift and int(shift.split('-')[-1].split(':')[0].strip(' ap')) < 12):
			day = True
		elif int(shift.split('-')[0].split(':')[0].strip(' ap')) < 5:
			day = True
		else:
			day = False
		return day
		
	def isEvening(shift):	
		
		evening = False
		
		if int(shift.split('-')[-1].split(':')[0].strip(' ap')) > 6:
			evening = True
		else:
			evening = False
			
		return evening
	
	
	dpop = week[weekday]
	eveningdpop = week[weekday + ' Evening']
	
	
	allShifts = sortData(weekday)
	
	eptPoint = allShifts[0]
	eptPointShifts = allShifts[1]
	
	ept = allShifts[2]
	eptShifts = allShifts[3]
	
	theaters = allShifts[4]
	theaterShifts = allShifts[5]
	
	lks = allShifts[6]
	lksShifts = allShifts[7]
	
	lksPoint = allShifts[8]
	lksPointShifts = allShifts[9]
	
	animals = allShifts[10]	
	animalsShifts = allShifts[11]


	def populateDpop(staff, shifts):
		
		start = 3
		eveStart = 3
		
		addNext = 0
		
		if staff == theaters:
			start = 25
			eveStart = 25
		elif staff == lks:
			start = 30
			eveStart = 30
		elif staff == lksPoint:
			start = 39
			eveStart = 39
		elif staff == animals:
			start = 42
			eveStart = 42
		
		
		
		if staff == ept:
			for i in range(0,10):
				if dpop.cell(row=21, column=start).value != None:
					start += 1
				if eveningdpop.cell(row=21, column=eveStart).value != None:
					eveStart += 1
		addEve = eveStart
		addDay = start
	
		for i in range(start, start+len(shifts)):
			if (isDay(shifts[addNext]) == True):
				dpop.cell(row=21, column = addDay).value = staff[addNext]
				dpop.cell(row=22, column = addDay).value = shifts[addNext]
				addDay += 1
			if isEvening(shifts[addNext]) == True:
				eveningdpop.cell(row=21, column=addEve).value = staff[addNext]
				eveningdpop.cell(row=22, column=addEve).value = shifts[addNext]
				addEve += 1
			addNext += 1
	

	
	
	populateDpop(eptPoint, eptPointShifts)
	populateDpop(ept, eptShifts)
	populateDpop(theaters, theaterShifts)
	populateDpop(lks, lksShifts)
	populateDpop(lksPoint, lksPointShifts)
	populateDpop(animals, animalsShifts)

	def trimDayDpops(staff):
		start = 3
		
		if staff == theaters:
			start = 25
		elif staff == lks:
			start = 30
		elif staff == lksPoint:
			start = 39
		elif staff == animals:
			start = 42
		
		end = start
		eveEnd = start
		
		for i in range(start, start+30):
			if dpop.cell(row=22, column=i).value != None:
				end += 1
			else:
				break
		
		for i in range(start, start+30):
			if eveningdpop.cell(row=22, column=i).value != None:
				eveEnd += 1
			else:
				break
				
		for h in range(start, end):
			shift = dpop.cell(row=22, column = h).value.split()	
			trimShift = dayShiftCells[dayShiftTimes.index(shift[0]): dayShiftTimes.index(shift[-1])]
			for i in range(23,61):
				if i in trimShift:
					pass
				else:
					dpop.cell(row=i, column=h).fill = greyFill
					dpop.cell(row=i, column=h).border = borderGone
					dpop.cell(row=i, column=h).value = ''
		
		for i in range(start, end):
			dpop.cell(row=22, column=i).value = dpop.cell(row=22, column=i).value.split()[0][:-1] + ' ' + dpop.cell(row=22, column=i).value.split()[-1][:-1]
		
		try:
			for h in range(start, eveEnd):
				eveShift = eveningdpop.cell(row=22, column=h).value.split()
				eveTrim = eveningShiftCells[eveningShiftTimes.index(eveShift[0]): eveningShiftTimes.index(eveShift[-1])]	
				for i in range(23, 41):
					if i in eveTrim:
						pass
					else:
						eveningdpop.cell(row=i, column=h).fill = greyFill
						eveningdpop.cell(row=i, column=h).border = borderGone
						eveningdpop.cell(row=i, column=h).value = ''
		except:
			pass
			
	trimDayDpops(ept)
	trimDayDpops(theaters)
	trimDayDpops(lks)
	trimDayDpops(lksPoint)
	trimDayDpops(animals)

#Run the program and make the shells	
browser = webdriver.Chrome()
browser.get('http://wheniwork.com/login')
week = openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoShells\\AutoShellsMasterExample.xlsx')

saveAs = input('What should I save this week as?')

addShifts('Thursday')
addShifts('Friday')
addShifts('Saturday')
addShifts('Sunday')
addShifts('Monday')
addShifts('Tuesday')
addShifts('Wednesday')

week.save('C:\\Users\\JOsbor01\\Desktop\\AutoShells\\' + saveAs + '.xlsx')

print('Shells are Done')
	
