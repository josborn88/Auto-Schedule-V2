def buildWindRove():
	import config
	import openpyxl, random, getpass, time
	from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
	from openpyxl.cell import Cell
	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

			
	blankFill = PatternFill(start_color = "FFFFFFFF", end_color = 'FFFFFFFF', fill_type = 'solid')
	travelFill = PatternFill(start_color = "0099CCFF", end_color = '0099CCFF', fill_type = 'solid')
	bklFill = PatternFill(start_color = "00FFFF00", end_color = '00FFFF00', fill_type = 'solid')
	#define show times
	windtubeReset = [55,56]
	
#assign first windtubes
	tried = 100
	go = True
	while go == True:
		assigned = False
		while assigned == False:
			who = random.randint(config.floorStart, config.floorEnd-1)
			if (config.dpop.cell(row = windtubeReset[0], column = who).fill == blankFill) and (config.dpop.cell(row=windtubeReset[1], column=who).fill == blankFill):
				config.dpop.cell(row = windtubeReset[0], column = who).fill = travelFill
				config.dpop.cell(row = windtubeReset[0], column = who).value = 'Windtubes'
				config.dpop.cell(row = windtubeReset[0], column = who).border = borderGoneB
				config.dpop.cell(row = windtubeReset[1], column = who).fill = travelFill
				config.dpop.cell(row = windtubeReset[1], column = who).border = borderGoneT
				assigned = True
				go = False
				tried = 0
			else:
				tried -= 1
				#print('Chem'+str(tried))
				if tried <= 0:
					go = False
					break
		else:
			tried-=1
			#print('chem'+str(tried))
			if tried <= 0:
					go=False
					break
#assign second windtube
	triedTwo = 100
	goTwo=True
	while goTwo==True:
		assigned = False
		while assigned == False:
			who = random.randint(config.floorStart, config.floorEnd-1)
			if (config.dpop.cell(row = windtubeReset[0], column = who).fill == blankFill) and (config.dpop.cell(row=windtubeReset[1], column=who).fill == blankFill):
				config.dpop.cell(row = windtubeReset[0], column = who).fill = travelFill
				config.dpop.cell(row = windtubeReset[0], column = who).value = 'Windtubes'
				config.dpop.cell(row = windtubeReset[0], column = who).border = borderGoneB
				config.dpop.cell(row = windtubeReset[1], column = who).fill = travelFill
				config.dpop.cell(row = windtubeReset[1], column = who).border = borderGoneT
				assigned = True
				goTwo = False
				triedTwo = 0
			else:
				tried -= 1
				#print('Chem'+str(tried))
				if tried <= 0:
					go = False
					break
		else:
			tried -= 1
			#print('chem'+str(tried))
			if tried <= 0:
					go = False
					break
					
