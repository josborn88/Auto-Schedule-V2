def buildMorningShow():
	import config
	import openpyxl, random, getpass, time
	from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
	from openpyxl.cell import Cell
	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	
	
	showdb=data['Show Training']	
	dbEnd=0
	for i in range(1,100):
		if showdb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#column numbers for workers
	lksWorkers=[]
	lksStart=0
	lksEnd=0
	#get the start of LKS workers
	for i in range(1,40):
		if config.dpop.cell(row=20,column=i).value=='little kidspace':
			lksStart=i
			break
	#get the end of lks workers
	for i in range(lksStart,50):
		if config.dpop.cell(row=21,column=i).value=='Name':
			lksEnd=i
			break
				
				
	for i in range(lksStart, lksEnd):
		lksWorkers.append(i)
	
	lksWorkers.append(lksEnd+1)

	#define station backgound colors
	blankFill= PatternFill(start_color="FFFFFFFF", end_color = 'FFFFFFFF', fill_type='solid')
	ratFill= PatternFill(start_color="FFCC00", end_color = 'FFCC00', fill_type='solid')
	bklFill= PatternFill(start_color="00FFFF00", end_color = '00FFFF00', fill_type='solid')
	
	
#get who is trained for shows
	knowsRatBall=[]
	knowsRatBallFirst=[]
	for i in range(1,dbEnd):
		if showdb.cell(row=i, column=82).value==True:
			knowsRatBall.append((showdb.cell(row=i,column=2).value.split()[0])+' '+showdb.cell(row=i,column=2).value.split()[1][0])
	
	for i in range(0,len(knowsRatBall)):
		knowsRatBallFirst.append(knowsRatBall[i].split()[0])
			
#define show times
	lksRatTime=[31,32,33,34]
	exploreTime=[31,32,33,34,35,36]
	animalTime=[43,44,45,46]
	storytime=[51,52]
#who knows explore more
	knowsExplore=[]
	knowsExploreFirst=[]
	for i in range(1,dbEnd):
		if showdb.cell(row=i, column=82).value==True:
			knowsExplore.append(showdb.cell(row=i,column=2).value.split()[0])
	
	for i in range(0,len(knowsExplore)):
		knowsExploreFirst.append(knowsExplore[i].split()[0])
			
#define show times
	exploreTime=[31,32,33,34,35,36]	
	
	
	if ('Thurs' in config.getday.sheetnames[0] or 'Sun' in config.getday.sheetnames[0]):
		#assign explore more
		tried = 100
		go=True
		while go==True:
			assigned=False
			while assigned==False:
				who=random.choice(lksWorkers)
				if ((config.dpop.cell(row=21,column=who).value in knowsExplore) or (config.dpop.cell(row=21,column=who).value in knowsExploreFirst))  and (config.dpop.cell(row=exploreTime[0], column=who).fill == blankFill) and (config.dpop.cell(row=exploreTime[2], column=who).fill == blankFill) and (config.dpop.cell(row=exploreTime[-1], column=who).fill == blankFill):
					for i in exploreTime:
						config.dpop.cell(row=i,column=who).fill=ratFill
						config.dpop.cell(row=exploreTime[0],column=who).fill=bklFill
						config.dpop.cell(row=exploreTime[0], column=who).value='Show Sheet'
						config.dpop.cell(row=exploreTime[0], column=who).border= borderGoneB
						config.dpop.cell(row=exploreTime[1],column=who).fill=ratFill
						config.dpop.cell(row=exploreTime[1],column=who).value='Prep'
						config.dpop.cell(row=exploreTime[1], column=who).border= borderGoneTB
						config.dpop.cell(row=exploreTime[2],column=who).fill=ratFill
						config.dpop.cell(row=exploreTime[2],column=who).value='LKS Exp.'
						config.dpop.cell(row=exploreTime[2], column=who).border= borderGoneTB
						config.dpop.cell(row=exploreTime[3],column=who).fill=ratFill
						config.dpop.cell(row=exploreTime[3],column=who).value='More'
						config.dpop.cell(row=exploreTime[3], column=who).border= borderGoneTBold
						config.dpop.cell(row=exploreTime[4],column=who).fill=ratFill
						config.dpop.cell(row=exploreTime[4],column=who).value='Clean Up'
						config.dpop.cell(row=exploreTime[4], column=who).border= borderGoneBBold
						config.dpop.cell(row=exploreTime[5],column=who).fill=ratFill
						config.dpop.cell(row=exploreTime[5], column=who).border= borderGoneT
						assigned=True
						for i in range(lksStart,lksEnd):
							if config.dpop.cell(row=exploreTime[-1], column=i).fill==ratFill:
								tried=0
								
				else:
					tried-=1
					#print('Chem'+str(tried))
					if tried<=0:
						go=False
						break
			else:
				tried-=1
				#print('chem'+str(tried))
				if tried <= 0:
						go=False
						break
	
	else:	
#assign lks rat ball
		tried = 100
		go=True
		while go==True:
			assigned=False
			while assigned==False:
				who=random.choice(lksWorkers)
				if ((config.dpop.cell(row=21,column=who).value in knowsRatBall) or (config.dpop.cell(row=21,column=who).value in knowsRatBallFirst)) and (config.dpop.cell(row=lksRatTime[0], column=who).fill == blankFill) and (config.dpop.cell(row=lksRatTime[2], column=who).fill == blankFill) and (config.dpop.cell(row=lksRatTime[3], column=who).fill == blankFill):
					config.dpop.cell(row=lksRatTime[0],column=who).fill=bklFill
					config.dpop.cell(row=lksRatTime[0], column=who).value='Show Sheet'
					config.dpop.cell(row=lksRatTime[0], column=who).border= borderGoneB
					config.dpop.cell(row=lksRatTime[1],column=who).value='Prep Rats'
					config.dpop.cell(row=lksRatTime[1],column=who).fill=ratFill
					config.dpop.cell(row=lksRatTime[1], column=who).border= borderGoneTB
					config.dpop.cell(row=lksRatTime[2],column=who).value='LKS Rat'
					config.dpop.cell(row=lksRatTime[2],column=who).fill=ratFill
					config.dpop.cell(row=lksRatTime[2], column=who).border= borderGoneTB
					config.dpop.cell(row=lksRatTime[3],column=who).fill=ratFill
					config.dpop.cell(row=lksRatTime[3],column=who).value='Bball'
					config.dpop.cell(row=lksRatTime[3], column=who).border= borderGoneTBold
					assigned=True
					if config.dpop.cell(row=lksRatTime[-1], column=who).fill==ratFill:
						tried=0
							
				else:
					tried-=1
					#print('Chem'+str(tried))
					if tried<=0:
						go=False
						break
			else:
				tried-=1
				#print('chem'+str(tried))
				if tried <= 0:
						go=False
						break
		