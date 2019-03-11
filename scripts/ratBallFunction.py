def buildRatBall():
	import config
	import openpyxl, random, getpass, time
	from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
	from openpyxl.cell import Cell
	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	showdb=data['Show Training']
	
	dbEnd=0
	for i in range(1,100):
		if showdb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#column numbers for workers
	floorStart=0
	floorEnd=0
	#get the start of Floor workers
	for i in range(1,40):
		if config.dpop.cell(row=20,column=i).value=='Paid Team (On-Floor)':
			floorStart=i
			break
	#get the end of workers
	for i in range(floorStart,40):
		if config.dpop.cell(row=21,column=i).value=='Name':
			floorEnd=i
			break
	
	
	#define station backgound colors
	blankFill= PatternFill(start_color="FFFFFFFF", end_color = 'FFFFFFFF', fill_type='solid')
	ratFill= PatternFill(start_color="FFCC00", end_color = 'FFCC00', fill_type='solid')
	bklFill= PatternFill(start_color="00FFFF00", end_color = '00FFFF00', fill_type='solid')	
	travelFill= PatternFill(start_color="0099CCFF", end_color = '0099CCFF', fill_type='solid')
#assign Rat Basketball
	knowsRat=[]
#get who is trained for rat basketball
	for i in range(1,dbEnd):
		if showdb.cell(row=i, column=51).value==True:
			knowsRat.append(showdb.cell(row=i,column=2).value.split()[0]+' '+showdb.cell(row=i,column=2).value.split()[1][0])

	knowsRatFirstName=[]
	for i in range(0,len(knowsRat)):
		knowsRatFirstName.append(knowsRat[i].split()[0])	
			
#define show times
	amRatBall=[39,40,41,42,43]
	pmRatBall=[51, 52, 53, 54, 55]

	tried = 100
	go=True
	while go==True:
		assigned=False
		while assigned==False:
			who=random.randint(floorStart,floorEnd-1)
			if ((config.dpop.cell(row=21,column=who).value in knowsRat) or (config.dpop.cell(row=21,column=who).value in knowsRatFirstName)) and (config.dpop.cell(row=amRatBall[0], column=who).fill == blankFill) and (config.dpop.cell(row=amRatBall[2], column=who).fill == blankFill) and (config.dpop.cell(row=amRatBall[4], column=who).fill == blankFill):
				for i in amRatBall:
					config.dpop.cell(row=i,column=who).fill=ratFill
					config.dpop.cell(row=44,column=who).fill=bklFill
					config.dpop.cell(row=amRatBall[0], column=who).value='Rats'
					config.dpop.cell(row=amRatBall[0], column=who).border = borderGoneBBold
					config.dpop.cell(row=amRatBall[1],column=who).value='Prep'
					config.dpop.cell(row=amRatBall[1], column=who).border = borderGoneTB
					config.dpop.cell(row=amRatBall[2],column=who).value='Rat B-Ball'
					config.dpop.cell(row=amRatBall[2], column=who).border = borderGoneTB
					config.dpop.cell(row=amRatBall[3],column=who).value='(EE)'
					config.dpop.cell(row=amRatBall[3], column=who).border = borderGoneTBold
					config.dpop.cell(row=amRatBall[4],column=who).value='Clean Up'
					config.dpop.cell(row=amRatBall[4], column=who).border = borderGoneBBold
					config.dpop.cell(row=44,column=who).value='Rover L2'
					assigned=True
					go = False
			else:
				tried-=1
				#print('rat'+str(tried))
				if tried<=0:
					go=False
					break
		else:
			tried-=1
			#print('Rat'+str(tried))
			if tried <= 0:
					go=False
					break
#assign second rat basketball	
	tried2 = 100
	go2=True
	while go2==True:
		assigned=False
		while assigned==False:
			who=random.randint(floorStart,floorEnd-1)
			if ((config.dpop.cell(row=21,column=who).value in knowsRat) or (config.dpop.cell(row=21,column=who).value in knowsRatFirstName)) and (config.dpop.cell(row=pmRatBall[0], column=who).fill == blankFill) and (config.dpop.cell(row=pmRatBall[2], column=who).fill == blankFill) and (config.dpop.cell(row=pmRatBall[4], column=who).fill == blankFill):
				for i in pmRatBall:
					config.dpop.cell(row=i,column=who).fill=ratFill
					config.dpop.cell(row=56, column=who).fill=travelFill
					config.dpop.cell(row=pmRatBall[0], column=who).value='Rats'
					config.dpop.cell(row=pmRatBall[0], column=who).border = borderGoneBBold
					config.dpop.cell(row=pmRatBall[1],column=who).value='Prep'
					config.dpop.cell(row=pmRatBall[1], column=who).border = borderGoneTB
					config.dpop.cell(row=pmRatBall[2],column=who).value='Rat B-Ball'
					config.dpop.cell(row=pmRatBall[2], column=who).border = borderGoneTB
					config.dpop.cell(row=pmRatBall[3],column=who).value='(EE)'
					config.dpop.cell(row=pmRatBall[3], column=who).border = borderGoneTBold
					config.dpop.cell(row=pmRatBall[4],column=who).value='Clean Up'
					config.dpop.cell(row=56,column=who).value='Windtubes'
					assigned=True
					go2 = False
			else:
				tried2-=1
				#print('ratPM'+str(tried2))
				if tried2<=0:
					go2=False
					break
		else:
			tried2-=1
			#print('RatPM'+str(tried2))
			if tried2 <= 0:
				go2=False
				break