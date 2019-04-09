import config, openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
from openpyxl.cell import Cell

class Shift:
    
	def __init__(self, name, clearance, long, restrictions, color, database, rowKnow):
		self.name = name
		self.clearance = clearance
		self.long = long
		self.restrictions = restrictions
		self.backFill = PatternFill(start_color = color, end_color = color, fill_type = 'solid')
		self.database = database
		self.rowKnow = rowKnow
		
class CloseShift:
	def __init__(self, name, color, colorTwo, database, rowKnow):
		self.name = name
		self.backFill = PatternFill(start_color = color, end_color = color, fill_type = 'solid')
		self.backFillTwo = PatternFill(start_color = colorTwo, end_color = colorTwo, fill_type = 'solid')
		self.database = database
		self.rowKnow = rowKnow
		
class ShowShift:
	def __init__(self, name, color, shiftStart, shiftEnd, database, rowKnow):
		self.name = name
		self.backFill = PatternFill(start_color = color, end_color = color, fill_type = 'solid')
		self.start = shiftStart
		self.end = shiftEnd
		self.database = database
		self.rowKnow = rowKnow
		
class OpenShift:
	def __init__(self, name, clearance, color, colorTwo, database, rowKnow, start, end):
		self.name = name
		self.clearance = clearance
		self.backFill = PatternFill(start_color = color, end_color = color, fill_type = 'solid')
		self.backFillTwo = PatternFill(start_color = colorTwo, end_color = colorTwo, fill_type = 'solid')
		self.database = database
		self.rowKnow = rowKnow
		self.start = start
		self.end = end