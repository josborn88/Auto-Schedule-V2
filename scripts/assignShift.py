import defineShifts, config, random, checkDatabase
from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold

def buildShift(shift):
	
	knows = checkDatabase.checkTraining(shift)
	knowsFirst = []
	for i in range(0, len(knows)):
		knowsFirst.append(knows[i].split()[0])
		
	
	tried = 1000
	go = True
	
	while go == True:
		for h in range(35,59, shift.long):
			assigned = False
			if shift.name == 'Cafe Host':
				if h > 51:
					go = False
					tried = 0
					assigned = True
			while assigned == False:
				who = random.randint(config.floorStart, config.floorEnd - 1)
				if (config.dpop.cell(row = h, column = who).fill == config.blankFill) and (config.dpop.cell(row = h + shift.long - 1, column = who).fill == config.blankFill):			
					if (config.dpop.cell(row = h - shift.long, column = who).value not in shift.restrictions) and ((config.dpop.cell(row=21,column=who).value in knows) or (config.dpop.cell(row=21,column=who).value in knowsFirst)):
						config.dpop.cell(row = h, column = who).value = shift.name.split()[0]
						if shift.name.split()[0] != shift.name.split()[-1]:
							config.dpop.cell(row = h + 1, column = who).value = shift.name.split()[1]
						config.dpop.cell(row = h, column = who).fill = shift.backFill
						config.dpop.cell(row = h + 1, column = who).fill = shift.backFill
						config.dpop.cell(row = h, column = who).border = config.borderGoneB
						config.dpop.cell(row = h + 1, column = who).border = config.borderGoneT
						assigned = True
						
						if shift.long == 4:
							config.dpop.cell(row = h + 2, column = who).fill = shift.backFill
							config.dpop.cell(row = h + 3, column = who).fill = shift.backFill
							config.dpop.cell(row = h + 1, column = who).border = config.borderGoneTB
							config.dpop.cell(row = h + 2, column = who).border = config.borderGoneTB
							config.dpop.cell(row = h + 3, column = who).border = config.borderGoneT
						for i in range(config.floorStart, config.floorEnd - 1):
							if config.dpop.cell(row = 58, column = i).fill == shift.backFill:
								tried = 0
								break
						
					else:
						tried -= 1
						if tried <= 0:
							go = False
							break
				else:
					tried -= 1
					if tried <= 0:
						go = False
						break	
		if tried <= 0:
			go = False
			break

def buildShiftExtended(shift):
	
	knows = checkDatabase.checkTraining(shift)
	knowsFirst = []
	for i in range(0, len(knows)):
		knowsFirst.append(knows[i].split()[0])
		
	
	tried = 1000
	go = True
	
	while go == True:
		for h in range(35,67, shift.long):
			assigned = False
			if shift.name == 'Cafe Host':
				if h > 56:
					go = False
					tried = 0
					assigned = True
			while assigned == False:
				who = random.randint(config.floorStart, config.floorEnd - 1)
				if (config.dpop.cell(row = h, column = who).fill == config.blankFill) and (config.dpop.cell(row = h + shift.long - 1, column = who).fill == config.blankFill):			
					if (config.dpop.cell(row = h - shift.long, column = who).value not in shift.restrictions) and ((config.dpop.cell(row=21,column=who).value in knows) or (config.dpop.cell(row=21,column=who).value in knowsFirst)):
						config.dpop.cell(row = h, column = who).value = shift.name.split()[0]
						if shift.name.split()[0] != shift.name.split()[-1]:
							config.dpop.cell(row = h + 1, column = who).value = shift.name.split()[1]
						config.dpop.cell(row = h, column = who).fill = shift.backFill
						config.dpop.cell(row = h + 1, column = who).fill = shift.backFill
						config.dpop.cell(row = h, column = who).border = config.borderGoneB
						config.dpop.cell(row = h + 1, column = who).border = config.borderGoneT
						assigned = True
						
						if shift.long == 4:
							config.dpop.cell(row = h + 2, column = who).fill = shift.backFill
							config.dpop.cell(row = h + 3, column = who).fill = shift.backFill
							config.dpop.cell(row = h + 1, column = who).border = config.borderGoneTB
							config.dpop.cell(row = h + 2, column = who).border = config.borderGoneTB
							config.dpop.cell(row = h + 3, column = who).border = config.borderGoneT
						for i in range(config.floorStart, config.floorEnd - 1):
							if config.dpop.cell(row = 66, column = i).fill == shift.backFill:
								tried = 0
								break
						
					else:
						tried -= 1
						if tried <= 0:
							go = False
							break
				else:
					tried -= 1
					if tried <= 0:
						go = False
						break	
		if tried <= 0:
			go = False
			break

		
def buildLksShift(shift):
	from openpyxl.styles import Color, colors, Font
	
	deskFont=Font(color=colors.WHITE)
	
	knows = checkDatabase.checkTraining(shift)
	knowsFirst = []
	for i in range(0, len(knows)):
		knowsFirst.append(knows[i].split()[0])
		
	
	tried = 1000
	go = True
	
	while go == True:
		for h in range(31, 59, shift.long):
			assigned = False
			while assigned == False:
				who = random.randint(config.lksStart, config.lksEnd - 1)
				if (config.dpop.cell(row = h, column = who).fill == config.blankFill) and (config.dpop.cell(row = h + shift.long - 1, column = who).fill == config.blankFill):			
					if (config.dpop.cell(row = h - shift.long, column = who).value not in shift.restrictions) and ((config.dpop.cell(row=21,column=who).value in knows) or (config.dpop.cell(row=21,column=who).value in knowsFirst)):
						config.dpop.cell(row = h, column = who).value = shift.name.split()[0]
						if shift.name.split()[0] != shift.name.split()[-1]:
							config.dpop.cell(row = h + 1, column = who).value = shift.name.split()[1]
						config.dpop.cell(row = h, column = who).fill = shift.backFill
						config.dpop.cell(row = h + 1, column = who).fill = shift.backFill
						config.dpop.cell(row = h + 2, column = who).fill = shift.backFill
						config.dpop.cell(row = h + 3, column = who).fill = shift.backFill
						if shift.name == 'LKS Desk':
							config.dpop.cell(row = i, column = who).font = deskFont
							config.dpop.cell(row = i + 1,column = who).font = deskFont
						
						assigned = True
						for i in range(config.lksStart, config.lksEnd - 1):
							if config.dpop.cell(row = 58, column = i).fill == shift.backFill:
								tried = 0
								break
						
					else:
						tried -= 1
						if tried <= 0:
							go = False
							break
				else:
					tried -= 1
					if tried <= 0:
						go = False
						break	
		if tried <= 0:
			go = False
			break

def buildShow(shift):
	
	times = {'10a': 31, '10:15a': 32, '10:30a': 33, '10:45a': 34, '11a': 35, '11:15a': 36, '11:30a': 37, '11:45a': 38, '12p': 39, '12:15p': 40, '12:30p':41, '12:45p':42, '1p':43, '1:15p': 44, '1:30p': 45, '1:45p': 46, '2p': 47, '2:15p': 48, '2:30p': 49, '2:45p': 50, '3p': 51, '3:15p': 52, '3:30p': 53, '3:45p': 54, '4p': 55, '4:15p': 56, '4:30p': 57, '4:45p': 58, '5p': 59, '5:15p': 60, '5:30p': 61}           
	
	tried = 1000
	go = True
	
	knows = checkDatabase.checkTraining(shift)
	knowsFirst = []
	
	for i in range(0, len(knows)):
		knowsFirst.append(knows[i].split()[0])

	while go == True:
		assigned = False
		while assigned == False:
			who = random.randint(config.floorStart, config.floorEnd - 1)
			if (config.dpop.cell(row = times.get(shift.start), column = who).fill == config.blankFill) and (config.dpop.cell(row = times.get(shift.end), column = who).fill == config.blankFill):			
				if ((config.dpop.cell(row=21,column=who).value in knows) or (config.dpop.cell(row=21,column=who).value in knowsFirst)):
					
					config.dpop.cell(row = times.get(shift.start), column = who).value = shift.name.split()[0]
					if shift.name.split()[0] != shift.name.split()[-1]:
							config.dpop.cell(row = times.get(shift.start) + 1, column = who).value = shift.name.split()[1]
					
					for h in range(times.get(shift.start), times.get(shift.end)):
						config.dpop.cell(row = h, column = who).fill = shift.backFill
						
					assigned = True
					tried = 0
					go = False
					
				else:
					tried -= 1
					if tried <= 0:
						go = False
						break
			else:
				tried -= 1
				if tried <= 0:
					go = False
					break	
	
					
					
def buildClose(shift):
	tried = 1000
	go = True
	
	knows = checkDatabase.checkTraining(shift)
	knowsFirst = []
	
	for i in range(0, len(knows)):
		knowsFirst.append(knows[i].split()[0])
		
	while go == True:	
		assigned = False
		while assigned == False:
			who = random.randint(config.floorStart, config.floorEnd - 1)
			if (config.dpop.cell(row=57, column=who).fill == config.blankFill) and (config.dpop.cell(row=58, column=who).fill == config.blankFill):			
				if ((config.dpop.cell(row=21,column=who).value in knows) or (config.dpop.cell(row=21,column=who).value in knowsFirst)):
					config.dpop.cell(row = 57, column = who).value = 'CL ' + shift.name.split()[0]
					if shift.name.split()[0] != shift.name.split()[-1]:
						config.dpop.cell(row = 58, column = who).value = 'CL ' + shift.name.split()[1]
					config.dpop.cell(row = 57, column = who).fill = shift.backFill
					config.dpop.cell(row = 58, column = who).fill = shift.backFillTwo
					assigned = True
					
					for i in range(config.floorStart, config.floorEnd - 1):
						if config.dpop.cell(row = 58, column = i).fill == shift.backFillTwo:
							tried = 0
							go = False
							break
					
				else:
					tried -= 1
					if tried <= 0:
						go = False
						break
			else:
				tried -= 1
				if tried <= 0:
					go = False
					break	
