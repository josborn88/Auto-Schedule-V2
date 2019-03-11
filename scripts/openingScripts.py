import config, random, checkDatabase, openpyxl
import openpyxl, random, getpass, time
from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
from openpyxl.cell import Cell
from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

def openHwu():

	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold

	hwuFill = config.PatternFill(start_color="0000FF00", end_color = '0000FF00', fill_type='solid')
	
	openHWU = []
	for i in range(1,50):
		if config.hwudb.cell(row=i, column=3).value == True:
			openHWU.append(config.hwudb.cell(row=i,column=2).value.split()[0] + ' ' + config.hwudb.cell(row=i,column=2).value.split()[1][0])
	
	openHWUFirstName = []
	
	for i in range(0, len(openHWU)):
		openHWUFirstName.append(openHWU[i].split()[0])
#assign HWU
	triedOpen = 1000
	goOpen = True
	while goOpen == True:
		
		assigned = False
		while assigned == False:
			who = random.randint(config.floorStart,config.floorEnd-1)
			if config.dpop.cell(row=31, column=who).fill == config.blankFill and (config.dpop.cell(row=33, column=who).fill == config.blankFill) and ((config.dpop.cell(row=21,column=who).value in openHWU) or (config.dpop.cell(row=21,column=who).value in openHWUFirstName)):
				
				config.dpop.cell(row=31, column=who).value = 'HWU'
				config.dpop.cell(row=32, column=who).value = 'Open'
				config.dpop.cell(row=33, column=who).value = '(Mezz)'
				
				config.dpop.cell(row=31, column=who).fill = hwuFill
				config.dpop.cell(row=32, column=who).fill = hwuFill
				config.dpop.cell(row=33, column=who).fill = hwuFill
				config.dpop.cell(row=34, column=who).fill = hwuFill
				
				config.dpop.cell(row=31, column=who).border = borderGoneB
				config.dpop.cell(row=32, column=who).border = borderGoneTB
				config.dpop.cell(row=33, column=who).border = borderGoneTB
				config.dpop.cell(row=34, column=who).border = borderGoneTBold
				assigned = True
				for i in range(config.floorStart,config.floorEnd):			
					if config.dpop.cell(row=31, column=i).value == 'HWU':
						triedOpen=0
				else:
					triedOpen -= 1
					#print('HWU'+str(tried))
					if triedOpen <= 0:
						goOpen=False
						break
			else:
				triedOpen -= 1
				#print('HWU'+str(tried))
				if triedOpen <= 0:
						goOpen = False
						break
		if triedOpen <= 0:
			goOpen = False
			break


def openOcean():

	dbEnd = 0
	for i in range(1,100):
		if config.opsdb.cell(row = i,column = 2).value == None:
			dbEnd = i
			break
			
	#fill patterns
	oceanFill=config.PatternFill(start_color="FF99CC", end_color = 'FF99CC', fill_type='solid')
	
	#marked off in database
	canOpen=[]
	for i in range(1,dbEnd):
		if config.opsdb.cell(row=i, column=13).value==True:
			canOpen.append(config.opsdb.cell(row=i,column=2).value.split()[0] + ' ' + config.opsdb.cell(row=i,column=2).value.split()[1][0])
	
	canOpenFirstName=[]
	for i in range(0,len(canOpen)):
		canOpenFirstName.append(canOpen[i].split()[0])
			
	#assign opening Shift
	tried = 1000
	go = True
	while go == True:
		assigned = False
		while assigned == False:
			who = random.randint(config.floorStart, config.floorEnd - 1)
			if (config.dpop.cell(row = 31, column = who).fill == config.blankFill) and (config.dpop.cell(row=32, column=who).fill == config.blankFill) and ((config.dpop.cell(row=21,column=who).value in canOpen) or (config.dpop.cell(row=21,column=who).value in canOpenFirstName)):
				config.dpop.cell(row = 31, column = who).value = 'Op Ocean'
				config.dpop.cell(row = 31, column = who).fill = oceanFill
				config.dpop.cell(row = 32, column = who).value = '& LP'
				config.dpop.cell(row = 32, column = who).fill = oceanFill
				triedSpace=0
				assigned=True
				go = False
			else:
				tried-=1
				#print('HWU'+str(tried))
				if tried <= 0:
						go=False
						break
		if tried <= 0:
			go=False
			break

def openSpace():

	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	opsdb=data['Floor Operations']
	
	dbEnd=0
	for i in range(1,100):
		if opsdb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#fill patterns
	spaceFill=PatternFill(start_color="cc99ff", end_color = 'cc99ff', fill_type='solid')
	#marked off in database
	canOpenSpace=[]
	for i in range(1,dbEnd):
		if opsdb.cell(row=i, column=11).value==True:
			canOpenSpace.append(opsdb.cell(row=i,column=2).value.split()[0] + ' ' + opsdb.cell(row=i,column=2).value.split()[1][0])
	
	canOpenSpaceFirstName=[]
	for i in range(0,len(canOpenSpace)):
		canOpenSpaceFirstName.append(canOpenSpace[i].split()[0])
			
	#assign opening Space
	triedSpace=1000
	go=True
	while go==True:
		assigned=False
		while assigned==False:
			who=random.randint(config.floorStart, config.floorEnd-1)
			if (config.dpop.cell(row=31, column=who).fill == config.blankFill) and (config.dpop.cell(row=32, column=who).fill == config.blankFill) and ((config.dpop.cell(row=21,column=who).value in canOpenSpace) or (config.dpop.cell(row=21,column=who).value in canOpenSpaceFirstName)):
				config.dpop.cell(row=31, column=who).value='Op Mezz'
				config.dpop.cell(row=31, column=who).fill=spaceFill
				config.dpop.cell(row=32, column=who).value='& Honda'
				config.dpop.cell(row=32, column=who).fill=spaceFill
				triedSpace=0
				assigned=True
				for i in range(config.floorStart,config.floorEnd):			
					if config.dpop.cell(row=31, column=i).value =='Op Mezz':
						triedOpen=0
					else:
						triedSpace-=1
						#print('HWU'+str(tried))
						if triedSpace <= 0:
							go=False
							break
			else:
				triedSpace-=1
				#print('HWU'+str(tried))
				if triedSpace <= 0:
						go=False
						break
		if triedSpace <= 0:
			go=False
			break


def openLife():

	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	opsdb=data['Floor Operations']
	
	dbEnd=0
	for i in range(1,100):
		if opsdb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#column numbers for workers
	floorStart=0
	floorEnd=0
	#get the start of Floor workers
	for i in range(1,40):
		if config.dpop.cell(row=20,column=i).value=='Paid Team (On-Floor)':
			floorStart=i
			break
	#get the end of workers
	for i in range(floorStart,40):
		if config.dpop.cell(row=21,column=i).value=='Name':
			floorEnd=i
			break
	#fill patterns
	progressFill= PatternFill(start_color="FFCC00", end_color = 'FFCC00', fill_type='solid')
	lifeFill= PatternFill(start_color="ff6600", end_color = 'ff6600', fill_type='solid')
	#marked off in database
	canOpenLifeProg=[]
	for i in range(1,dbEnd):
		if (opsdb.cell(row=i, column=7).value==True and opsdb.cell(row=i, column=9).value==True):
			canOpenLifeProg.append(opsdb.cell(row=i,column=2).value.split()[0] + ' ' + opsdb.cell(row=i,column=2).value.split()[1][0])
	
	canOpenLifeProgFirstName=[]
	for i in range(0,len(canOpenLifeProg)):
		canOpenLifeProgFirstName.append(canOpenLifeProg[i].split()[0])
			
	#assign opening Space
	tried=1000
	go=True
	while go==True:
		assigned=False
		while assigned==False:
			who=random.randint(config.floorStart, config.floorEnd-1)
			if (config.dpop.cell(row=31, column=who).fill == config.blankFill) and (config.dpop.cell(row=32, column=who).fill == config.blankFill) and ((config.dpop.cell(row=21,column=who).value in canOpenLifeProg) or (config.dpop.cell(row=21,column=who).value in canOpenLifeProgFirstName)):
				config.dpop.cell(row=31, column=who).value='Op Life'
				config.dpop.cell(row=31, column=who).fill=lifeFill
				config.dpop.cell(row=32, column=who).value='& Prog'
				config.dpop.cell(row=32, column=who).fill=progressFill
				tried=0
				assigned=True
				for i in range(config.floorStart, config.floorEnd):			
					if config.dpop.cell(row=31, column=i).value =='Op Life':
						triedOpen=0
					else:
						tried-=1
						#print('HWU'+str(tried))
						if tried <= 0:
							go=False
							break
			else:
				tried-=1
				#print('HWU'+str(tried))
				if tried <= 0:
						go=False
						break
		if tried <= 0:
			go=False
			break

def openGadgets():

	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold
	
	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	opsdb=data['Floor Operations']
	
	dbEnd=0
	for i in range(1,100):
		if opsdb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#fill patterns
	gadgetsFill= PatternFill(start_color="99cc00", end_color = '99cc00', fill_type='solid')
	cafeFill= PatternFill(start_color="E6B8B7", end_color = 'E6B8B7', fill_type='solid')
	#marked off in database
	canOpenGadgets=[]
	for i in range(1,dbEnd):
		if opsdb.cell(row=i, column=5).value==True:
			canOpenGadgets.append(opsdb.cell(row=i,column=2).value.split()[0] + ' ' + opsdb.cell(row=i,column=2).value.split()[1][0])
	
	canOpenGadgetsFirstName=[]
	for i in range(0,len(canOpenGadgets)):
		canOpenGadgetsFirstName.append(canOpenGadgets[i].split()[0])
			
	#assign opening Shift
	tried=1000
	go=True
	while go==True:
		assigned=False
		while assigned==False:
			who=random.randint(config.floorStart, config.floorEnd-1)
			if (config.dpop.cell(row=31, column=who).fill == config.blankFill) and (config.dpop.cell(row=32, column=who).fill == config.blankFill) and (config.dpop.cell(row=33, column=who).fill == config.blankFill) and (config.dpop.cell(row=34, column=who).fill == config.blankFill) and ((config.dpop.cell(row=21,column=who).value in canOpenGadgets) or (config.dpop.cell(row=21,column=who).value in canOpenGadgetsFirstName)):
				config.dpop.cell(row=31, column=who).value='Op Gad'
				config.dpop.cell(row=31, column=who).fill=gadgetsFill
				config.dpop.cell(row=32, column=who).value='Cafe'
				config.dpop.cell(row=32, column=who).fill=cafeFill
				config.dpop.cell(row=32, column=who).border=borderGoneTB
				config.dpop.cell(row=33, column=who).value='Prep'
				config.dpop.cell(row=33, column=who).fill=cafeFill
				config.dpop.cell(row=33, column=who).border=borderGoneTB
				config.dpop.cell(row=34, column=who).value='(Gad)'
				config.dpop.cell(row=34, column=who).fill=cafeFill
				config.dpop.cell(row=34, column=who).border=borderGoneTBold
				triedSpace=0
				assigned=True
				for i in range(config.floorStart, config.floorEnd):			
					if config.dpop.cell(row=31, column=i).value =='Op Mezz':
						tried=0
					else:
						tried-=1
						#print('HWU'+str(tried))
						if tried <= 0:
							go=False
							break
			else:
				tried-=1
				#print('HWU'+str(tried))
				if tried <= 0:
						go=False
						break
		if tried <= 0:
			go=False
			break
			
def openDinos():	

	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold
	
	dbEnd=0
	for i in range(1,100):
		if config.dinosdb.cell(row=i,column=2).value == None:
			dbEnd=i
			break
	
	dinoFill = config.PatternFill(start_color = "D8E4BC", end_color = 'D8E4BC', fill_type = 'solid')
#get who is trained on Dinos, talk to Aislinn regarding database format, if manager walkthrough means ok to do
	knowsDinos = []
	for i in range(2,dbEnd):
		if config.dinosdb.cell(row = i, column = 11).value != None:
			knowsDinos.append(config.dinosdb.cell(row=i,column=2).value.split()[0] + ' ' + config.dinosdb.cell(row=i,column=2).value.split()[1][0])
	
	knowsDinosFirstName=[]
	
	for i in range(0, len(knowsDinos)):
		knowsDinosFirstName.append(knowsDinos[i].split()[0])
#assign Dinos
	triedOpen = 1000
	goOpen = True
	while goOpen == True:
		assigned = False
		while assigned == False:
			who=random.randint(config.floorStart + 2,config.floorEnd - 1)
			if config.dpop.cell(row = 30, column = who).value == 'Meeting' and (config.dpop.cell(row = 31, column = who).fill == config.blankFill) and (config.dpop.cell(row = 33, column = who).fill == config.blankFill) and (config.dpop.cell(row = 34, column = who).fill == config.blankFill) and ((config.dpop.cell(row = 21,column = who).value in knowsDinos) or (config.dpop.cell(row = 21,column = who).value in knowsDinosFirstName)):
				
				config.dpop.cell(row = 30, column = who).value = 'Dinosaurs'
				config.dpop.cell(row = 31, column = who).value = 'Dinosaurs'
				config.dpop.cell(row = 30, column = who).fill = dinoFill
				config.dpop.cell(row = 31, column = who).fill = dinoFill
				config.dpop.cell(row = 32, column = who).fill = dinoFill
				config.dpop.cell(row = 33, column = who).fill = dinoFill
				config.dpop.cell(row = 34, column = who).fill = dinoFill
				
				config.dpop.cell(row = 30, column = who).border = borderGoneTBold
				config.dpop.cell(row = 31, column = who).border = borderGoneBBold
				config.dpop.cell(row = 32, column = who).border = borderGoneTB
				config.dpop.cell(row = 33, column = who).border = borderGoneTB
				config.dpop.cell(row = 34, column = who).border = borderGoneTBold
				
				assigned = True
				for i in range(config.floorStart,config.floorEnd):			
					if config.dpop.cell(row = 30, column = i).value =='Dinosaurs':
						triedOpen=0
				else:
					triedOpen -= 1
					#print('HWU'+str(tried))
					if triedOpen <= 0:
						goOpen=False
						break
			else:
				triedOpen -= 1
				#print('HWU'+str(tried))
				if triedOpen <= 0:
						goOpen = False
						break
		if triedOpen <= 0:
			goOpen = False
			break
			
def openAmnh():

	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

	dbEnd=0
	for i in range(1,100):
		if config.tempdb.cell(row = i,column = 2).value == None:
			dbEnd = i
			break
			
	
	amnhFill = config.PatternFill(start_color = "00339966", end_color = '00339966', fill_type = 'solid')
#check database
	knowsAmnh = []
	for i in range(2,dbEnd):
		if config.tempdb.cell(row = i, column = 9).value != None:
			knowsAmnh.append(config.tempdb.cell(row = i,column = 2).value.split()[0] + ' ' + config.tempdb.cell(row=i,column=2).value.split()[1][0])
	
	knowsAmnhFirstName = []
	
	for i in range(0,len(knowsAmnh)):
		knowsAmnhFirstName.append(knowsAmnh[i].split()[0])
#assign amnh
	triedOpen = 1000
	goOpen = True
	while goOpen == True:
		assigned = False
		while assigned == False:
			who = random.randint(config.floorStart + 2,config.floorEnd - 1)
			if config.dpop.cell(row = 30, column = who).value == 'Meeting' and (config.dpop.cell(row = 31, column = who).fill == config.blankFill) and (config.dpop.cell(row=33, column=who).fill == config.blankFill) and (config.dpop.cell(row=34, column=who).fill == config.blankFill) and ((config.dpop.cell(row=21,column=who).value in knowsAmnh) or (config.dpop.cell(row=21,column=who).value in knowsAmnhFirstName)):
				config.dpop.cell(row = 30, column = who).value = 'Mythic'
				config.dpop.cell(row = 31, column = who).value = 'Creatures'
				config.dpop.cell(row = 30, column = who).fill = amnhFill
				config.dpop.cell(row = 31, column = who).fill = amnhFill
				config.dpop.cell(row = 32, column = who).fill = amnhFill
				config.dpop.cell(row = 33, column = who).fill = amnhFill
				config.dpop.cell(row = 34, column = who).fill = amnhFill
				
				config.dpop.cell(row = 30, column = who).border = borderGoneTBold
				config.dpop.cell(row = 31, column = who).border = borderGoneBBold
				config.dpop.cell(row = 32, column = who).border = borderGoneTB
				config.dpop.cell(row = 33, column = who).border = borderGoneTB
				config.dpop.cell(row = 34, column = who).border = borderGoneTBold
				
				assigned = True
				for i in range(config.floorStart,config.floorEnd):			
					if config.dpop.cell(row = 30, column = i).value == 'Mythic':
						triedOpen = 0
				else:
					triedOpen -= 1
					#print('HWU'+str(tried))
					if triedOpen <= 0:
						goOpen = False
						break
			else:
				triedOpen -= 1
				#print('HWU'+str(tried))
				if triedOpen <= 0:
						goOpen = False
						break
		if triedOpen <= 0:
			goOpen = False
			break

def openAmnhTwo():

	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

	dbEnd=0
	for i in range(1,100):
		if config.tempdb.cell(row = i,column = 2).value == None:
			dbEnd = i
			break
			
	
	amnhFill = config.PatternFill(start_color = "00339966", end_color = '00339966', fill_type = 'solid')
#check database
	knowsAmnh = []
	for i in range(2,dbEnd):
		if config.tempdb.cell(row = i, column = 9).value != None:
			knowsAmnh.append(config.tempdb.cell(row = i,column = 2).value.split()[0] + ' ' + config.tempdb.cell(row=i,column=2).value.split()[1][0])
	
	knowsAmnhFirstName = []
	
	for i in range(0,len(knowsAmnh)):
		knowsAmnhFirstName.append(knowsAmnh[i].split()[0])
#assign amnh
	triedOpen = 1000
	goOpen = True
	while goOpen == True:
		assigned = False
		while assigned == False:
			who = random.randint(config.floorStart + 2,config.floorEnd - 1)
			if config.dpop.cell(row = 30, column = who).value == 'Meeting' and (config.dpop.cell(row = 31, column = who).fill == config.blankFill) and (config.dpop.cell(row=33, column=who).fill == config.blankFill) and (config.dpop.cell(row=34, column=who).fill == config.blankFill) and ((config.dpop.cell(row=21,column=who).value in knowsAmnh) or (config.dpop.cell(row=21,column=who).value in knowsAmnhFirstName)):
				config.dpop.cell(row = 30, column = who).value = 'Mythic'
				config.dpop.cell(row = 31, column = who).value = 'Creatures'
				config.dpop.cell(row = 30, column = who).fill = amnhFill
				config.dpop.cell(row = 31, column = who).fill = amnhFill
				config.dpop.cell(row = 32, column = who).fill = amnhFill
				config.dpop.cell(row = 33, column = who).fill = amnhFill
				config.dpop.cell(row = 34, column = who).fill = amnhFill
				
				config.dpop.cell(row = 30, column = who).border = borderGoneTBold
				config.dpop.cell(row = 31, column = who).border = borderGoneBBold
				config.dpop.cell(row = 32, column = who).border = borderGoneTB
				config.dpop.cell(row = 33, column = who).border = borderGoneTB
				config.dpop.cell(row = 34, column = who).border = borderGoneTBold
				
				assigned = True
				for i in range(config.floorStart,config.floorEnd):			
					if config.dpop.cell(row = 30, column = i).value == 'Mythic':
						triedOpen = 0
				else:
					triedOpen -= 1
					#print('HWU'+str(tried))
					if triedOpen <= 0:
						goOpen = False
						break
			else:
				triedOpen -= 1
				#print('HWU'+str(tried))
				if triedOpen <= 0:
						goOpen = False
						break
		if triedOpen <= 0:
			goOpen = False
			break