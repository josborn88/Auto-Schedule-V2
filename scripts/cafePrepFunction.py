def buildCafePrep():
	import config
	import openpyxl, random, getpass, time
	from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
	from openpyxl.cell import Cell
	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	cafedb=data['Gadgets Cafe']
	
	dbEnd=0
	for i in range(1,100):
		if cafedb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#column numbers for workers
	floorStart=0
	floorEnd=0
	#get the start of Floor workers
	for i in range(1,40):
		if config.dpop.cell(row=20,column=i).value=='Paid Team (On-Floor)':
			floorStart=i
			break
	#get the end of workers
	for i in range(floorStart,40):
		if config.dpop.cell(row=21,column=i).value=='Name':
			floorEnd=i
			break
			
	blankFill= PatternFill(start_color="FFFFFFFF", end_color = 'FFFFFFFF', fill_type='solid')
	cafeFill= PatternFill(start_color="E6B8B7", end_color = 'E6B8B7', fill_type='solid')
#check database
	knowsCafePrep=[]
	for i in range(1,dbEnd):
		if cafedb.cell(row=i, column=8).value==True:
			knowsCafePrep.append(cafedb.cell(row=i,column=2).value.split()[0]+' '+cafedb.cell(row=i,column=2).value.split()[1][0])	
	
	knowsCafePrepFirst=[]
	for i in range(0,len(knowsCafePrep)):
		knowsCafePrepFirst.append(knowsCafePrep[i].split()[0])		
#define show times
	cafePrepTime=[55,56]
#assign first Cafe prep
	tried = 100
	go=True
	while go==True:
		assigned=False
		while assigned==False:
			who=random.randint(floorStart,floorEnd-1)
			if ((config.dpop.cell(row=21,column=who).value in knowsCafePrep) or (config.dpop.cell(row=21,column=who).value in knowsCafePrepFirst)) and (config.dpop.cell(row=cafePrepTime[0], column=who).fill == blankFill) and (config.dpop.cell(row=cafePrepTime[1], column=who).fill == blankFill):
				config.dpop.cell(row=cafePrepTime[0],column=who).fill=cafeFill
				config.dpop.cell(row=cafePrepTime[0], column=who).value='Cafe Prep'
				config.dpop.cell(row=cafePrepTime[0], column=who).border= borderGoneB
				config.dpop.cell(row=cafePrepTime[1],column=who).fill=cafeFill
				config.dpop.cell(row=cafePrepTime[1], column=who).value='(G)'
				config.dpop.cell(row=cafePrepTime[1], column=who).border= borderGoneT
				assigned=True
				if config.dpop.cell(row=cafePrepTime[-1], column=who).fill==cafeFill:
					tried=0
						
			else:
				tried-=1
				#print('Chem'+str(tried))
				if tried<=0:
					go=False
					break
		else:
			tried-=1
			#print('chem'+str(tried))
			if tried <= 0:
					go=False
					break
#assign second Cafe prep
	triedTwo = 100
	goTwo=True
	while goTwo==True:
		assignedTwo=False
		while assignedTwo==False:
			who=random.randint(floorStart,floorEnd-1)
			if ((config.dpop.cell(row=21,column=who).value in knowsCafePrep) or (config.dpop.cell(row=21,column=who).value in knowsCafePrepFirst)) and (config.dpop.cell(row=cafePrepTime[0], column=who).fill == blankFill) and (config.dpop.cell(row=cafePrepTime[1], column=who).fill == blankFill):
				config.dpop.cell(row=cafePrepTime[0],column=who).fill=cafeFill
				config.dpop.cell(row=cafePrepTime[0], column=who).value='Cafe Prep'
				config.dpop.cell(row=cafePrepTime[0], column=who).border= borderGoneB
				config.dpop.cell(row=cafePrepTime[1],column=who).fill=cafeFill
				config.dpop.cell(row=cafePrepTime[1], column=who).value='(P&L)'
				config.dpop.cell(row=cafePrepTime[1], column=who).border= borderGoneT
				assignedTwo=True
				if config.dpop.cell(row=cafePrepTime[-1], column=who).fill==cafeFill:
					triedTwo=0
						
			else:
				triedTwo-=1
				#print('Chem'+str(tried))
				if triedTwo<=0:
					goTwo=False
					break
		else:
			triedTwo-=1
			#print('chem'+str(tried))
			if triedTwo <= 0:
					goTwo=False
					break