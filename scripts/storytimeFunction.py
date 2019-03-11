def buildStorytime():
	import config
	import openpyxl, random, getpass, time
	from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
	from openpyxl.cell import Cell
	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	
	
	showdb=data['Show Training']	
	dbEnd=0
	for i in range(1,100):
		if showdb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#column numbers for workers
	floorStart=0
	floorEnd=0
	#get the start of Floor workers
	for i in range(1,40):
		if config.dpop.cell(row=20,column=i).value=='Paid Team (On-Floor)':
			floorStart=i
			break
	#get the end of workers
	for i in range(floorStart,40):
		if config.dpop.cell(row=21,column=i).value=='Name':
			floorEnd=i
			break
	lksStart=0
	lksEnd=0
	#get the start of LKS workers
	for i in range(1,40):
		if config.dpop.cell(row=20,column=i).value=='little kidspace':
			lksStart=i
			break
	#get the end of lks workers
	for i in range(lksStart,50):
		if config.dpop.cell(row=21,column=i).value=='Name':
			lksEnd=i
			break		
#fill patterns			
	blankFill= PatternFill(start_color="FFFFFFFF", end_color = 'FFFFFFFF', fill_type='solid')
	ratFill= PatternFill(start_color="FFCC00", end_color = 'FFCC00', fill_type='solid')
	bklFill= PatternFill(start_color="00FFFF00", end_color = '00FFFF00', fill_type='solid')
#get who is trained for shows
	knowsStory=[]
	knowsStoryFirst=[]
	for i in range(1,dbEnd):
		if showdb.cell(row=i, column=79).value==True:
			knowsStory.append((showdb.cell(row=i,column=2).value.split()[0])+' '+showdb.cell(row=i,column=2).value.split()[1][0])
	
		for i in range(0,len(knowsStory)):
			knowsStoryFirst.append(knowsStory[i].split()[0])		
#define show times
	storytime=[51,52]
		
#assign lks storytime
	tried = 100
	go=True
	while go==True:
		assigned=False
		while assigned==False:
			who=random.randint(lksStart,lksEnd-1)
			if ((config.dpop.cell(row=21,column=who).value in knowsStory) or (config.dpop.cell(row=21,column=who).value in knowsStoryFirst)) and (config.dpop.cell(row=storytime[0], column=who).fill == blankFill) and (config.dpop.cell(row=storytime[1], column=who).fill == blankFill):
				config.dpop.cell(row=storytime[0],column=who).fill=ratFill
				config.dpop.cell(row=storytime[0], column=who).value='Storytime'
				config.dpop.cell(row=storytime[0], column=who).border= borderGoneB
				config.dpop.cell(row=storytime[1],column=who).fill=ratFill
				config.dpop.cell(row=storytime[1], column=who).border= borderGoneT
				assigned=True
				if config.dpop.cell(row=storytime[-1], column=who).fill==ratFill:
					tried=0
						
			else:
				tried-=1
				#print('Chem'+str(tried))
				if tried<=0:
					go=False
					break
		else:
			tried-=1
			#print('chem'+str(tried))
			if tried <= 0:
					go=False
					break