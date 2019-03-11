def buildEG():
	import config
	import openpyxl, random, getpass, time
	from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
	from openpyxl.cell import Cell
	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	
	showdb=data['Show Training']	
	dbEnd=0
	for i in range(1,100):
		if showdb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#column numbers for workers
	floorStart=0
	floorEnd=0
	#get the start of Floor workers
	for i in range(1,40):
		if config.dpop.cell(row=20,column=i).value=='Paid Team (On-Floor)':
			floorStart=i
			break
	#get the end of workers
	for i in range(floorStart,40):
		if config.dpop.cell(row=21,column=i).value=='Name':
			floorEnd=i
			break

	#define station backgound colors
	blankFill= PatternFill(start_color="FFFFFFFF", end_color = 'FFFFFFFF', fill_type='solid')
	egFill= PatternFill(start_color="00CCFFFF", end_color = '00CCFFFF', fill_type='solid')
		
	
#Check Database for who is trained
	knowsEG=[]
	
	for i in range(1,dbEnd):
		if showdb.cell(row=i, column=59).value==True:
			knowsEG.append(showdb.cell(row=i,column=2).value.split()[0])
	
#define show times
	firstEG=[33,34]
	secondEG=[45,46]
	lastEG=[57,58]
	
#assign first EG
	tried = 100
	go=True
	while go==True:
		assigned=False
		while assigned==False:
			who=random.randint(floorStart,floorEnd-1)
			if (config.dpop.cell(row=21,column=who).value.split()[0] in knowsEG) and (config.dpop.cell(row=firstEG[0], column=who).fill == blankFill):
				for i in firstEG:
					config.dpop.cell(row=i,column=who).fill=egFill
					config.dpop.cell(row=firstEG[0], column=who).value='EG'
					config.dpop.cell(row=firstEG[0], column=who).border=borderGoneTB
					config.dpop.cell(row=firstEG[1],column=who).value='(EE)'
					config.dpop.cell(row=firstEG[1], column=who).border=borderGoneTB
					assigned=True
					for i in range(floorStart,floorEnd):
						if config.dpop.cell(row=firstEG[-1], column=i).fill ==egFill:
							tried=0
			else:
				tried-=1
				#print('EG'+str(tried))
				if tried<=0:
					go=False
					break
		else:
			tried-=1
			#print('EG'+str(tried))
			if tried <= 0:
					go=False
					break
#assign second EG	
	tried2 = 100
	go2=True
	while go2==True:
		assigned=False
		while assigned==False:
			who=random.randint(floorStart,floorEnd-1)
			if (config.dpop.cell(row=21,column=who).value.split()[0] in knowsEG) and (config.dpop.cell(row=secondEG[0], column=who).fill == blankFill):
				for i in secondEG:
					config.dpop.cell(row=i,column=who).fill=egFill
					config.dpop.cell(row=secondEG[0], column=who).value='EG'
					config.dpop.cell(row=secondEG[0], column=who).border= borderGoneTB
					config.dpop.cell(row=secondEG[1],column=who).value='(EE)'
					config.dpop.cell(row=secondEG[1], column=who).border= borderGoneTB
					assigned=True
					for i in range(floorStart,floorEnd):
						if config.dpop.cell(row=secondEG[-1], column=i).fill ==egFill:
							tried2=0
			else:
				tried2-=1
				#print('EG2'+str(tried2))
				if tried2<=0:
					go2=False
					break
		else:
			tried2-=1
			#print('EG2'+str(tried2))
			if tried2 <= 0:
				go2=False
				break

#assign last EG	
	tried3 = 100
	go3=True
	while go3==True:
		assigned=False
		while assigned==False:
			who=random.randint(floorStart,floorEnd-1)
			if (config.dpop.cell(row=21,column=who).value.split()[0] in knowsEG) and (config.dpop.cell(row=lastEG[0], column=who).fill == blankFill):
				for i in lastEG:
					config.dpop.cell(row=i,column=who).fill=egFill
					config.dpop.cell(row=lastEG[0], column=who).value='EG'
					config.dpop.cell(row=lastEG[0], column=who).border=borderGoneTB
					config.dpop.cell(row=lastEG[1],column=who).value='(EE)'
					config.dpop.cell(row=lastEG[1], column=who).border=borderGoneTB
					assigned=True
					for i in range(floorStart,floorEnd):
						if config.dpop.cell(row=lastEG[-1], column=i).fill ==egFill:
							tried3=0
			else:
				tried3-=1
				#print('EG2'+str(tried3))
				if tried3<=0:
					go3=False
					break
		else:
			tried2-=1
			#print('EG2'+str(tried3))
			if tried3 <= 0:
				go3=False
				break