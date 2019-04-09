import openpyxl, random, getpass, time, os
from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
from openpyxl.cell import Cell

dbPath = os.path.dirname(os.path.dirname( __file__ ))
#what is the seasonal gadgets stage?
#season=input('what is the seasonal Gadgets show?').split()
season='Chem Live'
#Load Database
data=openpyxl.load_workbook(dbPath + '\\floorDatabase.xlsx')
showdb=data['Show Training']
hwudb=data['Attractions']
cafedb=data['Gadgets Cafe']
tempdb=data['Temporary Programs']
dinosdb=data['Dinos']
lksdb=data['LKS Duties']
opsdb=data['Floor Operations']	
		
borderGoneTB = Border(left=Side(border_style='thin', color='000000'),
				right=Side(border_style='thin', color='000000'),
				top=Side(border_style='none', color='000000'),
				bottom=Side(border_style='none', color='000000'))
				
borderGoneB = Border(left=Side(border_style='thin', color='000000'),
				right=Side(border_style='thin', color='000000'),
				top=Side(border_style='thin', color='000000'),
				bottom=Side(border_style='none', color='000000'))

borderGoneT = Border(left=Side(border_style='thin', color='000000'),
				right=Side(border_style='thin', color='000000'),
				top=Side(border_style='none', color='000000'),
				bottom=Side(border_style='thin', color='000000'))

borderGoneTBold = Border(left=Side(border_style='thin', color='000000'),
				right=Side(border_style='thin', color='000000'),
				top=Side(border_style='none', color='000000'),
				bottom=Side(border_style='medium', color='000000'))

borderGoneBBold = Border(left=Side(border_style='thin', color='000000'),
				right=Side(border_style='thin', color='000000'),
				top=Side(border_style='medium', color='000000'),
				bottom=Side(border_style='none', color='000000'))

greyFill = PatternFill(start_color="C0C0C0", end_color = 'C0C0C0', fill_type='solid')
blankFill = PatternFill(start_color = "FFFFFFFF", end_color = 'FFFFFFFF', fill_type = 'solid')
meetingFill = PatternFill(start_color = '00FFFF00', end_color = '00FFFF00', fill_type = 'solid')
#set start and end of EPT colums

whatDay=input('What day are we making?')
getday=openpyxl.load_workbook(dbPath + '\\' + whatDay + '.xlsx')
dpop=getday[getday.sheetnames[0]]

floorStart=3
floorEnd=0
#get the start of Floor workers

#get the end of workers
for i in range(floorStart,100):
	if dpop.cell(row=21,column=i).value=='Name':
		floorEnd=i
		break

lksStart=0
lksEnd=0
	#get the start of LKS workers
for i in range(1,40):
	if dpop.cell(row=20,column=i).value=='little kidspace':
		lksStart=i
		break
#get the end of lks workers
for i in range(lksStart,50):
	if dpop.cell(row=21,column=i).value=='Name':
		lksEnd=i
		break		