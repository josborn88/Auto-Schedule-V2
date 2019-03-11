import openpyxl, config
from openpyxl.styles import Font
def eptTest():


#test the dpop to make sure critical things are made
	wb = openpyxl.load_workbook(config.dbPath + '\\process\\' + config.whatDay + ' EPT.xlsx')
	ws = wb[wb.sheetnames[0]]
	
	Dinos = 1
	Amnh = 2
	Travel = 1
	Cafe = 0
	HWU = 1
	
	for i in range(config.floorStart, config.floorEnd):
		for h in range(35, 59, 4):
			if ws.cell(row = h,column = i).value == "Dinosaurs":
				Dinos += 1
			else:
				continue
		
	for i in range(config.floorStart, config.floorEnd):
		for h in range(35, 59, 4):
			if ws.cell(row = h,column = i).value == "Mythic":
				Amnh += 1
			else:
				continue
	for i in range(config.floorStart, config.floorEnd):
		for h in range(35, 59, 4):
			if ws.cell(row = h,column = i).value == "Crocs":
				Travel += 1
			else:
				continue
	for i in range(config.floorStart, config.floorEnd):
		for h in range(35, 59, 2):
			if ws.cell(row = h,column = i).value == "HWU" or ws.cell(row = h,column = i).value == "HWU CL":
				HWU += 1
			else:
				continue
	for i in range(config.floorStart, config.floorEnd):
		for h in range(35, 55, 4):
			if ws.cell(row = h,column = i).value == "Cafe":
				Cafe += 1
			else:
				continue
				
	#Check the sums of the shift to make sure everything is assigned correctly
	if Dinos == 7 and Travel == 7 and HWU == 13 and Cafe == 5 and Amnh == 14:
		return True
	else:
		return False
		
def lksTest():

#test the dpop to make sure critical things are made
	wb = openpyxl.load_workbook(config.dbPath + '\\finished\\' + config.whatDay + '.xlsx')
	ws = wb[wb.sheetnames[0]]
	
	desk = 0
	bkl = 0
	studio = 0
	
	for i in range(config.lksStart, config.lksEnd):
		for h in range(31, 59, 4):
			if ws.cell(row = h + 1,column = i).value == "Desk":
				desk += 1
			else:
				continue
		
	for i in range(config.lksStart, config.lksEnd):
		for h in range(31, 59, 4):
			if ws.cell(row = h + 1, column = i).value == "BKL":
				bkl += 1
			else:
				continue
				
	for i in range(config.lksStart, config.lksEnd):
		for h in range(31, 59, 4):
			if ws.cell(row = h + 1, column = i).value == "Studio":
				studio += 1
			else:
				continue
	
				
	#Check the sums of the shift to make sure everything is assigned correctly
	if desk == 7 and bkl == 7 and studio == 7:
		return True
	else:
		return False