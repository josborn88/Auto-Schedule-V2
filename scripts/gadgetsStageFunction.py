def buildGadgetsStage():
	import config
	import openpyxl, random, getpass, time
	from openpyxl.styles import Color, PatternFill, Font, Border, Side, colors
	from openpyxl.cell import Cell
	from config import borderGoneTB, borderGoneB, borderGoneT, borderGoneTBold, borderGoneBBold, greyFill

	data=openpyxl.load_workbook('C:\\Users\\JOsbor01\\Desktop\\AutoPop\\floorDatabase.xlsx')
	
	season='Chem Live'
	
	showdb=data['Show Training']	
	dbEnd=0
	for i in range(1,100):
		if showdb.cell(row=i,column=2).value==None:
			dbEnd=i
			break
			
	#column numbers for workers
	floorStart=0
	floorEnd=0
	#get the start of Floor workers
	for i in range(1,40):
		if config.dpop.cell(row=20,column=i).value=='Paid Team (On-Floor)':
			floorStart=i
			break
	#get the end of workers
	for i in range(floorStart,40):
		if config.dpop.cell(row=21,column=i).value=='Name':
			floorEnd=i
			break
			
	#define station backgound colors
	blankFill= PatternFill(start_color="FFFFFFFF", end_color = 'FFFFFFFF', fill_type='solid')
	chemFill= PatternFill(start_color="0000CCFF", end_color = '0000CCFF', fill_type='solid')
	
	
	
#get who is trained for chem live
	knowsChem=[]
	knowsSeason=[]
	for i in range(1,dbEnd):
		if showdb.cell(row=i, column=8).value==True:
			knowsChem.append(showdb.cell(row=i,column=2).value.split()[0])
	
	for i in range(1,dbEnd):
		if season[0].lower()=='fireworks':
				if showdb.cell(row=i, column=25).value==True:
					knowsSeason.append(showdb.cell(row=i,column=2).value.split()[0])
		elif season[0].lower()=='pumpkin':
			if showdb.cell(row=i, column=17).value==True:
				knowsSeason.append(showdb.cell(row=i,column=2).value.split()[0])
		elif season[0].lower()=='so':
			if showdb.cell(row=i, column=33).value==True:
				knowsSeason.append(showdb.cell(row=i,column=2).value.split()[0])
		else:
			if showdb.cell(row=i, column=8).value==True:
				knowsSeason.append(showdb.cell(row=i,column=2).value.split()[0])
			
#define show times
	amChem=[33,34,35,36,37,38]
	pmChem=[45,46,47,48,49,50]
#assign am Chem Live
	tried = 100
	go=True
	while go==True:
		assigned=False
		while assigned==False:
			who=random.randint(floorStart,floorEnd-1)
			if (config.dpop.cell(row=21,column=who).value.split()[0] in knowsChem) and (config.dpop.cell(row=amChem[0], column=who).fill == blankFill) and (config.dpop.cell(row=amChem[2], column=who).fill == blankFill) and (config.dpop.cell(row=amChem[4], column=who).fill == blankFill) and (config.dpop.cell(row=amChem[-1], column=who).fill == blankFill):
				for i in amChem:
					config.dpop.cell(row=i,column=who).fill=chemFill
					config.dpop.cell(row=amChem[0], column=who).value='Chem'
					config.dpop.cell(row=amChem[0], column=who).border= borderGoneB
					config.dpop.cell(row=amChem[1],column=who).value='Live'
					config.dpop.cell(row=amChem[1], column=who).border= borderGoneTBold
					config.dpop.cell(row=amChem[2], column=who).border= borderGoneBBold
					config.dpop.cell(row=amChem[3], column=who).border= borderGoneTB
					config.dpop.cell(row=amChem[4], column=who).border= borderGoneTB
					config.dpop.cell(row=amChem[4], column=who).value= 'Clean'
					config.dpop.cell(row=amChem[5], column=who).border= borderGoneTBold
					config.dpop.cell(row=amChem[5], column=who).value= 'Up'
					assigned=True
					for i in range(floorStart,floorEnd):
						if config.dpop.cell(row=amChem[-1], column=i).fill==chemFill:
							tried=0
			else:
				tried-=1
				#print('Chem'+str(tried))
				if tried<=0:
					go=False
					break
		else:
			tried-=1
			#print('chem'+str(tried))
			if tried <= 0:
					go=False
					break
#assign pm Chem Live	

	tried2 = 100
	go2=True
	while go2==True:
		assigned=False
		while assigned==False:
			who=random.randint(floorStart,floorEnd-1)
			if (config.dpop.cell(row=21,column=who).value.split()[0] in knowsSeason) and (config.dpop.cell(row=pmChem[0], column=who).fill == blankFill) and (config.dpop.cell(row=pmChem[2], column=who).fill == blankFill) and (config.dpop.cell(row=pmChem[4], column=who).fill == blankFill) and (config.dpop.cell(row=pmChem[-1], column=who).fill == blankFill):
				for i in pmChem:
					config.dpop.cell(row=i,column=who).fill=chemFill
				config.dpop.cell(row=pmChem[0], column=who).value=season.split()[0]

				if len(season)>1:
					config.dpop.cell(row=pmChem[1],column=who).value=season.split()[1]
				assigned=True
				for i in range(floorStart,floorEnd):
					if config.dpop.cell(row=pmChem[-1], column=i).fill==chemFill:
						tried2=0
				config.dpop.cell(row=pmChem[0], column=who).border= borderGoneB
				config.dpop.cell(row=pmChem[1], column=who).border= borderGoneTBold
				config.dpop.cell(row=pmChem[2], column=who).border= borderGoneBBold
				config.dpop.cell(row=pmChem[3], column=who).border= borderGoneTB
				config.dpop.cell(row=pmChem[4], column=who).border= borderGoneTB
				config.dpop.cell(row=pmChem[4], column=who).value= 'Clean'
				config.dpop.cell(row=pmChem[5], column=who).border= borderGoneTBold
				config.dpop.cell(row=pmChem[5], column=who).value= 'Up'
				
			else:
				tried2-=1
				#print('ChemPM'+str(tried2))
				if tried2<=0:
					go2=False
					break
		else:
			tried2-=1
			#print('ChemPM'+str(tried2))
			if tried2 <= 0:
				go2=False
				break